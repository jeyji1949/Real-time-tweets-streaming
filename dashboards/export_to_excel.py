#!/usr/bin/env python3
"""
Exporte vers Excel avec formatage
"""

from elasticsearch import Elasticsearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference
from datetime import datetime

class ExcelExporter:
    def __init__(self):
        self.es = Elasticsearch(["http://localhost:9200"])
    
    def export_to_excel(self, filename=None):
        """Crée un fichier Excel avec plusieurs feuilles"""
        if filename is None:
            filename = f"twitter_analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        wb = Workbook()
        
        # Feuille 1 : Tous les tweets
        self._create_tweets_sheet(wb)
        
        # Feuille 2 : Stats par topic
        self._create_topic_stats_sheet(wb)
        
        # Feuille 3 : Top hashtags
        self._create_hashtags_sheet(wb)
        
        # Supprimer la feuille par défaut
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Sauvegarder
        wb.save(filename)
        print(f"✅ Excel créé : {filename}")
        
        return filename
    
    def _create_tweets_sheet(self, wb):
        """Feuille avec les tweets"""
        ws = wb.active
        ws.title = "Tweets"
        
        # Headers
        headers = ['ID', 'Texte', 'User', 'Sentiment', 'Confiance', 'Topic', 'Likes', 'RT']
        ws.append(headers)
        
        # Style des headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Récupérer les tweets
        result = self.es.search(
            index="tweets_index_improved",
            body={"query": {"match_all": {}}, "size": 1000}
        )
        
        # Ajouter les données
        for hit in result['hits']['hits']:
            tweet = hit['_source']
            ws.append([
                tweet.get('tweet_id', ''),
                tweet.get('text', ''),
                tweet.get('user', ''),
                tweet.get('sentiment', ''),
                tweet.get('confidence', 0),
                tweet.get('topic', ''),
                tweet.get('like_count', 0),
                tweet.get('retweet_count', 0)
            ])
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
    
    def _create_topic_stats_sheet(self, wb):
        """Feuille avec stats par topic"""
        ws = wb.create_sheet("Stats par Topic")
        
        # Headers
        headers = ['Topic', 'Nombre', 'Confiance Moy.', 'Likes Moy.', 'RT Moy.']
        ws.append(headers)
        
        # Style
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Query
        query = {
            "size": 0,
            "aggs": {
                "by_topic": {
                    "terms": {"field": "topic.keyword"},
                    "aggs": {
                        "avg_confidence": {"avg": {"field": "confidence"}},
                        "avg_likes": {"avg": {"field": "like_count"}},
                        "avg_rt": {"avg": {"field": "retweet_count"}}
                    }
                }
            }
        }
        
        result = self.es.search(index="tweets_index_improved", body=query)
        
        for bucket in result['aggregations']['by_topic']['buckets']:
            ws.append([
                bucket['key'],
                bucket['doc_count'],
                round(bucket['avg_confidence']['value'], 2),
                round(bucket['avg_likes']['value'], 1),
                round(bucket['avg_rt']['value'], 1)
            ])
        
        # Graphique (Pie chart)
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribution par Topic"
        
        ws.add_chart(pie, "G2")
    
    def _create_hashtags_sheet(self, wb):
        """Feuille avec top hashtags"""
        ws = wb.create_sheet("Top Hashtags")
        
        headers = ['Rang', 'Hashtag', 'Nombre']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Query
        query = {
            "size": 0,
            "aggs": {
                "top_hashtags": {
                    "terms": {"field": "hashtags.keyword", "size": 20}
                }
            }
        }
        
        result = self.es.search(index="tweets_index_improved", body=query)
        
        for i, bucket in enumerate(result['aggregations']['top_hashtags']['buckets'], 1):
            ws.append([i, bucket['key'], bucket['doc_count']])


if __name__ == "__main__":
    exporter = ExcelExporter()
    exporter.export_to_excel()